import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image as XLImage

HEADER_FILL = PatternFill(start_color="1F4E5F", end_color="1F4E5F", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="DCE9EC", end_color="DCE9EC", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
GRAND_TOTAL_FILL = PatternFill(start_color="1F4E5F", end_color="1F4E5F", fill_type="solid")

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("SR.NO", 6),
    ("DATE", 12),
    ("REASON", 30),
    ("MODE", 10),
    ("FROM", 16),
    ("TO", 16),
    ("OTHER", 12),
    ("CNG/BUS", 12),
    ("KM", 8),
    ("COURIER/\nTRANSPORT", 14),
    ("FOOD", 12),
    ("TOTAL", 14),
]


def _fmt_date(d):
    if d is None:
        return ""
    try:
        return d.strftime("%d-%m-%y")
    except AttributeError:
        return str(d)


def generate_expense_excel(output_path, company, employee, expenses, period_label,
                            report_title="EMPLOYEE EXPENSE SHEET"):
    """
    company: dict with name, address_line1, address_line2, mobile
    employee: dict with name, designation, employee_code
    expenses: list of expense-like objects/dicts with the expense fields
    period_label: e.g. "15 JUL 2026 TO 30 JUL 2026"
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Sheet"

    n_cols = len(COLUMNS)
    last_col_letter = get_column_letter(n_cols)

    # ---------- Header block ----------
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = company.get("name", "")
    ws["A1"].font = Font(size=15, bold=True, color="1F4E5F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"A2:{last_col_letter}2")
    addr = f"{company.get('address_line1', '')} {company.get('address_line2', '')}".strip()
    ws["A2"] = addr
    ws["A2"].font = Font(size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A3:{last_col_letter}3")
    ws["A3"] = f"Mobile: {company.get('mobile', '')}"
    ws["A3"].font = Font(size=10)
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A4:{last_col_letter}4")
    ws["A4"] = report_title
    ws["A4"].font = Font(size=13, bold=True, color="FFFFFF")
    ws["A4"].fill = HEADER_FILL
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 22

    # ---------- Employee info row ----------
    ws.merge_cells("A5:D5")
    ws["A5"] = f"Employee: {employee.get('name', '')}"
    ws.merge_cells("E5:H5")
    ws["E5"] = f"Designation: {employee.get('designation', '')}"
    ws.merge_cells(f"I5:{last_col_letter}5")
    ws["I5"] = f"Expense Period: {period_label}"
    for cell_ref in ("A5", "E5", "I5"):
        ws[cell_ref].font = Font(size=10, bold=True)
        ws[cell_ref].fill = SUBHEADER_FILL
        ws[cell_ref].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[5].height = 18

    ws.row_dimensions[6].height = 4  # spacer

    # ---------- Optional company logo (top-left, floats over the header) ----------
    logo_path = company.get("logo_path")
    if logo_path and os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            img.height = 50
            img.width = 50
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = max(ws.row_dimensions[1].height or 0, 38)
        except Exception:
            pass  # never let a bad logo file break the export

    # ---------- Table header ----------
    header_row = 7
    for idx, (col_name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[header_row].height = 30

    # ---------- Data rows ----------
    row = header_row + 1
    totals = {"other": 0, "cng_bus": 0, "km": 0, "courier": 0, "food": 0, "grand": 0}

    for sr, e in enumerate(expenses, start=1):
        values = [
            sr,
            _fmt_date(_get(e, "expense_date")),
            _get(e, "reason"),
            _get(e, "mode"),
            _get(e, "from_location"),
            _get(e, "to_location"),
            _get(e, "other_amount") or None,
            _get(e, "cng_bus_amount") or None,
            _get(e, "km") or None,
            _get(e, "courier_transport_amount") or None,
            _get(e, "food_amount") or None,
            _get(e, "total_amount") or 0,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx in (7, 8, 10, 11, 12) and isinstance(val, (int, float)):
                cell.number_format = '"₹"#,##0.00'
            if col_idx == 3:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        totals["other"] += _get(e, "other_amount") or 0
        totals["cng_bus"] += _get(e, "cng_bus_amount") or 0
        totals["km"] += _get(e, "km") or 0
        totals["courier"] += _get(e, "courier_transport_amount") or 0
        totals["food"] += _get(e, "food_amount") or 0
        totals["grand"] += _get(e, "total_amount") or 0
        row += 1

    if not expenses:
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")
        ws.cell(row=row, column=1, value="No expenses recorded for this period.")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        row += 1

    data_end_row = row - 1

    # ---------- Totals block ----------
    row += 1
    totals_start = row
    label_col = n_cols - 1
    value_col = n_cols

    total_lines = [
        ("TOTAL OTHER", totals["other"]),
        ("TOTAL CNG/BUS", totals["cng_bus"]),
        ("TOTAL KM", totals["km"]),
        ("TOTAL COURIER/TRANSPORT", totals["courier"]),
        ("TOTAL FOOD", totals["food"]),
    ]
    for label, val in total_lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=label_col - 1)
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(bold=True, size=10)
        c1.alignment = Alignment(horizontal="right", vertical="center")
        c1.fill = TOTAL_FILL

        c2 = ws.cell(row=row, column=label_col, value=val)
        c2.number_format = '"₹"#,##0.00'
        c2.font = Font(bold=True, size=10)
        c2.fill = TOTAL_FILL
        c2.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=row, start_column=label_col + 1, end_row=row, end_column=value_col)
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=label_col - 1)
    gc1 = ws.cell(row=row, column=1, value="GRAND TOTAL")
    gc1.font = Font(bold=True, size=12, color="FFFFFF")
    gc1.fill = GRAND_TOTAL_FILL
    gc1.alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells(start_row=row, start_column=label_col, end_row=row, end_column=value_col)
    gc2 = ws.cell(row=row, column=label_col, value=totals["grand"])
    gc2.number_format = '"₹"#,##0.00'
    gc2.font = Font(bold=True, size=12, color="FFFFFF")
    gc2.fill = GRAND_TOTAL_FILL
    gc2.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[row].height = 22

    footer_row = row + 2
    ws.merge_cells(f"A{footer_row}:{last_col_letter}{footer_row}")
    gen_note = ws.cell(
        row=footer_row, column=1,
        value=f"Generated by Siddhivinayak Expense Manager on {datetime.now().strftime('%d-%m-%Y %H:%M')}",
    )
    gen_note.font = Font(size=8, italic=True, color="808080")
    gen_note.alignment = Alignment(horizontal="center")

    # ---------- Sheet-level settings ----------
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{max(data_end_row, header_row)}"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)
    ws.print_area = f"A1:{last_col_letter}{footer_row}"
    ws.oddHeader.center.text = "&B" + report_title
    ws.oddFooter.center.text = "Page &P of &N"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def _get(obj, attr):
    if isinstance(obj, dict):
        return obj.get(attr)
    return getattr(obj, attr, None)
