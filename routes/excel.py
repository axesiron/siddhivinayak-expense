import os
from datetime import date
from flask import Blueprint, send_file, request, current_app, abort
from flask_login import login_required, current_user
from models import Expense, Employee, CompanySettings
from utils.excel_generator import generate_expense_excel

excel_bp = Blueprint("excel", __name__, url_prefix="/export/excel")


def get_company_dict():
    row = CompanySettings.query.first()
    if row and row.company_name:
        return {
            "name": row.company_name,
            "address_line1": row.address_line1,
            "address_line2": row.address_line2,
            "mobile": row.mobile,
            "logo_path": row.logo_path,
        }
    return {
        "name": current_app.config["COMPANY_NAME"],
        "address_line1": current_app.config["COMPANY_ADDRESS_LINE1"],
        "address_line2": current_app.config["COMPANY_ADDRESS_LINE2"],
        "mobile": current_app.config["COMPANY_MOBILE"],
        "logo_path": None,
    }


def build_and_send(expenses, employee, period_label, filename, report_title="EMPLOYEE EXPENSE SHEET"):
    company = get_company_dict()
    employee_dict = {
        "name": employee.name,
        "designation": employee.designation,
        "employee_code": employee.employee_code,
    }
    output_path = os.path.join(current_app.config["EXPORT_FOLDER"], filename)
    generate_expense_excel(output_path, company, employee_dict, expenses, period_label, report_title)
    return send_file(output_path, as_attachment=True, download_name=filename)


@excel_bp.route("/")
@login_required
def export_all():
    """All of the current employee's expenses."""
    items = Expense.query.filter_by(employee_id=current_user.id).order_by(Expense.expense_date.asc()).all()
    period_label = "ALL RECORDS"
    filename = f"{current_user.employee_code}_all_expenses.xlsx"
    return build_and_send(items, current_user, period_label, filename)


@excel_bp.route("/custom")
@login_required
def export_custom():
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")

    query = Expense.query.filter_by(employee_id=current_user.id)
    if from_date:
        query = query.filter(Expense.expense_date >= from_date)
    if to_date:
        query = query.filter(Expense.expense_date <= to_date)
    items = query.order_by(Expense.expense_date.asc()).all()

    period_label = f"{from_date or 'START'} TO {to_date or 'TODAY'}"
    filename = f"{current_user.employee_code}_{from_date}_to_{to_date}.xlsx"
    return build_and_send(items, current_user, period_label, filename)


@excel_bp.route("/monthly")
@login_required
def export_monthly():
    from sqlalchemy import extract
    month = int(request.args.get("month", date.today().month))
    year = int(request.args.get("year", date.today().year))

    items = Expense.query.filter(
        Expense.employee_id == current_user.id,
        extract("month", Expense.expense_date) == month,
        extract("year", Expense.expense_date) == year,
    ).order_by(Expense.expense_date.asc()).all()

    period_label = f"MONTH {month:02d}-{year}"
    filename = f"{current_user.employee_code}_{year}_{month:02d}.xlsx"
    return build_and_send(items, current_user, period_label, filename, report_title="MONTHLY EXPENSE REPORT")


@excel_bp.route("/employee")
@login_required
def export_employee():
    """Admin: export a specific employee's expense report."""
    if not current_user.is_admin:
        abort(403)

    emp_id = request.args.get("employee_id", type=int)
    employee = Employee.query.get_or_404(emp_id)
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")

    query = Expense.query.filter_by(employee_id=employee.id)
    if from_date:
        query = query.filter(Expense.expense_date >= from_date)
    if to_date:
        query = query.filter(Expense.expense_date <= to_date)
    items = query.order_by(Expense.expense_date.asc()).all()

    period_label = f"{from_date or 'START'} TO {to_date or 'TODAY'}"
    filename = f"{employee.employee_code}_report.xlsx"
    return build_and_send(items, employee, period_label, filename)


@excel_bp.route("/all-employees")
@login_required
def export_all_employees():
    """Admin: export every employee's expenses into one workbook (multi-sheet)."""
    if not current_user.is_admin:
        abort(403)

    from openpyxl import Workbook
    employees = Employee.query.filter_by(role="employee").order_by(Employee.name).all()
    company = get_company_dict()

    wb = Workbook()
    wb.remove(wb.active)

    from utils.excel_generator import generate_expense_excel
    import tempfile

    # Build a combined workbook by generating each employee sheet into a temp
    # file then copying data across (kept simple & robust for large exports).
    output_path = os.path.join(current_app.config["EXPORT_FOLDER"], "all_employees_expenses.xlsx")
    os.makedirs(current_app.config["EXPORT_FOLDER"], exist_ok=True)

    from openpyxl import load_workbook
    final_wb = Workbook()
    final_wb.remove(final_wb.active)

    for emp in employees:
        items = Expense.query.filter_by(employee_id=emp.id).order_by(Expense.expense_date.asc()).all()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        generate_expense_excel(tmp_path, company,
                                {"name": emp.name, "designation": emp.designation,
                                 "employee_code": emp.employee_code},
                                items, "ALL RECORDS")
        src_wb = load_workbook(tmp_path)
        src_ws = src_wb.active
        safe_title = (emp.name or emp.employee_code)[:28]
        new_ws = final_wb.create_sheet(title=safe_title)
        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.alignment = cell.alignment.copy()
                    new_cell.number_format = cell.number_format
        for col, dim in src_ws.column_dimensions.items():
            new_ws.column_dimensions[col].width = dim.width
        os.remove(tmp_path)

    final_wb.save(output_path)
    return send_file(output_path, as_attachment=True, download_name="all_employees_expenses.xlsx")
